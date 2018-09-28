from tkinter import *
from tkinter import ttk
from tkinter.filedialog import askopenfilename, asksaveasfile, asksaveasfilename
import datetime
from datetime import datetime, timedelta
import SCF_interface as inter
import sqlite3
import tkinter.font as tkFont
import tkinter.ttk as ttk
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Font

def criar_conexao():
	global con
	con = Conexao()

class Colaborador():
	def __init__(self, nome="", dtNasc="", lab="", funcao="", ch="", dtIngresso="", status="", cpf="", senha=""):
		self.nome = nome
		self.dtNasc = dtNasc
		self.lab = lab
		self.funcao = funcao
		self.ch = ch
		self.dtIngresso = dtIngresso
		self.status = status
		self.cpf = cpf
		self.senha = senha

class Conexao():
	def __init__(self):
		self.conexao = sqlite3.connect(r"\\Lsehost\scf\SCF.db")
		self.cursor = self.conexao.cursor()
		self.cursor.execute("CREATE TABLE IF NOT EXISTS Laboratorio (Nome VARCHAR(45), Sigla VARCHAR(45), Logo BLOB, Image_type VARCHAR(3), PRIMARY KEY(Nome, Sigla))")
		self.cursor.execute("CREATE TABLE IF NOT EXISTS Colaborador (Nome VARCHAR(45), DtNasc DATE, Lab VARCHAR(45), Funcao VARCHAR(45), CH INT, DtIngresso DATE, DtDesligamento DATE, Status VARCHAR(45), cpf VARCHAR(11), Foto BLOB, Image_type VARCHAR(3), Senha VARCHAR(45) NOT NULL, PRIMARY KEY(cpf))")
		self.cursor.execute("INSERT OR IGNORE INTO Laboratorio(Nome, Sigla) VALUES(?,?)", ("Administração", "ADM"))
		self.conexao.commit()


def cadastrar_colaborador(nome, DtNasc, Lab, Funcao, CH, DtIngresso, status, cpf, senha, confirma_senha, foto, event=None):
	if senha.get() != confirma_senha.get():
		inter.pop_up("ERROR", "Confirme sua Senha!")
	elif(not(validar_cpf(cpf.get()))):
		inter.pop_up("ERROR", "CPF inválido")
	elif(not(validar_data(DtNasc.get())) or not(validar_data(DtIngresso.get()))):
		inter.pop_up("ERROR", "Data inválida")
	else:
		cursor = con.cursor
		conexao = con.conexao

		try:
			if (Lab.get() != "*Selecione o laboratório*") or (Funcao.get() in ['ADM', 'Coordenador Geral']):

				lab = ""
				if(Funcao.get() in ['ADM', 'Coordenador Geral']):
					lab = "Administração"
				else:
					lab = Lab.get()

				if foto.get() != "":
					file_type = ImageMethods.get_file_type(foto.get())
					file_binary = ImageMethods.get_binary(foto.get())
					cursor.execute("INSERT INTO Colaborador (Nome, DtNasc, Lab, Funcao, CH, DtIngresso, Status, cpf, Foto, Image_type, Senha) VALUES(?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)", (nome.get(), DtNasc.get(), lab, Funcao.get(), CH.get(), DtIngresso.get(), status.get(), cpf.get(), file_binary, file_type, senha.get()))
				else:
					cursor.execute("INSERT INTO Colaborador (Nome, DtNasc, Lab, Funcao, CH, DtIngresso, Status, cpf, Foto, Image_type, Senha) VALUES(?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)", (nome.get(), DtNasc.get(), lab, Funcao.get(), CH.get(), DtIngresso.get(), status.get(), cpf.get(), None, None, senha.get()))
				conexao.commit()
				return True
			else:
				inter.pop_up("ERROR", "Cadastro Inválido!")
				return False	
		except Exception:
			inter.pop_up("ERROR", "Cadastro Inválido")
			return False
	return False

def atualizar_cadastro_colaborador(nome, DtNasc, Lab, Funcao, CH, DtIngresso, status, senha, confirma_senha, foto, cpf_colab):
	if senha.get()!=confirma_senha.get():
		inter.pop_up("ERROR", "Confirme sua Senha!")
	else:
		cursor = con.cursor
		conexao = con.conexao

		try:
			if foto.get() != "":
				file_type = ImageMethods.get_file_type(foto.get())
				file_binary = ImageMethods.get_binary(foto.get())
				cursor.execute('''UPDATE Colaborador 
					SET Nome = (?), DtNasc = (?), Lab = (?), Funcao=(?), CH=(?), DtIngresso=(?), Status=(?), Foto=(?), Image_type = (?), Senha=(?) WHERE cpf = (?) ''', (nome.get(), DtNasc.get(), Lab.get(), Funcao.get(), CH.get(), DtIngresso.get(), status.get(), file_binary, file_type, senha.get(), cpf_colab))
			else:
				cursor.execute('''UPDATE Colaborador 
					SET Nome=(?), DtNasc=(?), Lab=(?), Funcao=(?), CH=(?), DtIngresso=(?), Status=(?), Senha=(?) WHERE cpf = (?) ''', (nome.get(), DtNasc.get(), Lab.get(), Funcao.get(), CH.get(), DtIngresso.get(), status.get(), senha.get(), cpf_colab))
			conexao.commit()
			return True
		except:
			inter.pop_up("ERROR", "Cadastro Inválido")
			return False
	return False

def atualizar_cadastro_laboratorio(nome, sigla, logo_path):
	if logo_path.get() != "":
		cursor = con.cursor
		conexao = con.conexao
		try:
			file_type = ImageMethods.get_file_type(logo_path.get())
			file_binary = ImageMethods.get_binary(logo_path.get())
			cursor.execute('''UPDATE Laboratorio 
					SET Logo = (?), Image_type = (?) WHERE Nome = (?) and Sigla = (?)''', (file_binary, file_type, nome.get(), sigla.get()))
			conexao.commit()

			inter.pop_up("Sucesso", "Dados Atualizados")
			return True

		except Exception as e:
			print(e)
			inter.pop_up("ERROR", "O caminho informado é inválido")
			return False
	else:
		return False

def cadastrar_laboratorio(nome, sigla, logo):
	cursor = con.cursor
	conexao = con.conexao
	if nome.get() != "" and sigla.get()!="":
		try:
			if logo.get() != "":
				file_type = ImageMethods.get_file_type(logo.get())
				file_binary = ImageMethods.get_binary(logo.get())
				cursor.execute("INSERT INTO Laboratorio VALUES (?, ?, ?, ?)", (nome.get(), sigla.get(), file_binary, file_type))
			else:
				cursor.execute("INSERT INTO Laboratorio VALUES (?, ?, ?, ?)", (nome.get(), sigla.get(), None, None))

			conexao.commit()
			return True
		except:
			inter.pop_up("ERROR", "Cadastro Inválido")
	else:
		inter.pop_up("ERROR", "Cadastro Inválido")
	return False

def retorna_lista_lab():
	cursor = con.cursor
	cursor.execute('''SELECT Nome
				   FROM Laboratorio''')
	lista = []
	for nome in cursor.fetchall():
		lista.append(str(nome[0]))
	return lista

def retorna_lista_colab(lab):
	cursor = con.cursor
	if lab == 'Não Ativos':
		cursor.execute("SELECT Nome FROM Colaborador WHERE Status='Não Ativo'")
	else:
		cursor.execute("SELECT Nome FROM Colaborador WHERE Lab = '%s' and Status='Ativo' ORDER BY Nome"%lab)
	lista = []
	for nome in cursor.fetchall():
		lista.append(str(nome[0]))
	return lista

def retorna_dados_lab(lab):
	cursor = con.cursor
	try:	
		cursor.execute("SELECT Nome,Sigla FROM Laboratorio WHERE Nome='%s'"%(lab))
		return cursor.fetchall()[0]
	except:
		pop_up("ERROR", "Laboratorio não encontrado!")

def retorna_colab(nome, lab):
	cursor = con.cursor
	if type(nome) != str:
		if lab == 'Não Ativos':
			cursor.execute("SELECT * FROM Colaborador WHERE Nome = '%s' and Status = 'Não Ativo'"%nome.get())
		else:
			cursor.execute("SELECT * FROM Colaborador WHERE Nome = '%s' and Lab='%s'"%(nome.get(), lab))
	else:
		if lab == 'Não Ativos':
			cursor.execute("SELECT * FROM Colaborador WHERE Nome = '%s' and Status = 'Não Ativo'"%nome)
		else:
			cursor.execute("SELECT * FROM Colaborador WHERE Nome = '%s' and Lab='%s'"%(nome, lab))

	lista = cursor.fetchall()
	tupla = lista[0]
	colab = Colaborador(tupla[0], tupla[1], tupla[2], tupla[3], tupla[4], tupla[5], tupla[7], tupla[8], tupla[11])
	return colab


def retorna_user(cpf):
	cursor = con.cursor
	cursor.execute("SELECT * FROM Colaborador WHERE cpf = '%s'"%cpf)
	lista = cursor.fetchall()
	tupla = lista[0]
	colab = Colaborador(tupla[0], tupla[1], tupla[2], tupla[3], tupla[4], tupla[5], tupla[7], tupla[8], tupla[11])
	return colab

def retorna_objeto_date(data):
	data_hora = data.split(" ")
	data = data_hora[0].split("/")
	hora = data_hora[1].split(":")
	date = datetime(int(data[0]),int(data[1]),int(data[2]), int(hora[0]),int(hora[1]),int(hora[2]))
	return date

class ImageMethods():
	@staticmethod
	def get_path(line_path):
		tk_dialog = Tk()
		tk_dialog.withdraw()
		file_types=[('PNG file',"*.png"), ('JPG file', "*.jpg")]
		file_name = askopenfilename(filetypes=file_types, title="Selecione o logo")
		line_path.set(file_name)
		tk_dialog.destroy()

	@staticmethod
	def get_binary(image_path):
		binary = ''
		with open(image_path, 'rb') as image:
			binary = image.read()
		return binary

	@staticmethod
	def get_file_type(image_path):
		file_type = ""
		for i in range(-3,0):
			file_type = file_type + image_path[i]
		return  (file_type)

def retorna_data():
	now = datetime.now()
	
	dia = str(now.day)
	if len(dia)<2:
		dia = '0'+dia
	
	mes = str(now.month)
	if len(mes)<2:
		mes = '0'+mes
	
	ano = str(now.year)
	
	data = ano+'/'+mes+'/'+dia
	
	return data


def validar_cpf(cpf):
	try:
		if cpf.isnumeric() and (len(cpf) == 11):
			igual = True
			for i in range(1,11):
				if cpf[i] != cpf[i-1]:	
					igual = False
			if (igual):
				return False

			cpf_soma = 0

			for i in range(0, 9):
				cpf_soma = cpf_soma + int(cpf[i])*(10-i)

			if ((cpf_soma*10)%11 == int(cpf[9])) or ((cpf_soma*10)%11 == 10 and int(cpf[9]) == 0):
				cpf_soma = 0
				for i in range(0, 10):
					cpf_soma = cpf_soma + int(cpf[i])*(11-i)

				if ((cpf_soma*10)%11 == int(cpf[10])) or ((cpf_soma*10)%11 == 10 and int(cpf[10]) == 0):
					return True
				
			else:
				return False
		return False
	except:
		return False

def validar_data(data):
	try:
		datetime.strptime(data, '%d/%m/%Y')
		return True
	except Exception as e:
		print(e)
		return False

def validar_chamada_historico(lab, mes, ano, tipo):
	laboratorio, mes, ano = lab.get(), mes.get(), ano.get()

	if laboratorio != "*Selecione o laboratório*":
		try:
			cursor = con.cursor
			dict_mes = {'Janeiro': '01','Fevereiro': '02','Março': '03','Abril': '04','Maio': '05','Junho': '06',
						'Julho': '07','Agosto': '08','Setembro': '09','Outubro': '10','Novembro': '11','Dezembro': '12'}

			date = ano+"/"+dict_mes[mes]

			if laboratorio == "Não Ativos":
				cursor.execute("SELECT cpf,Nome From Colaborador WHERE Status = 'Não Ativo'")
			else:
				cursor.execute("SELECT cpf,Nome From Colaborador WHERE Lab = '%s' ORDER BY Nome"%laboratorio)

			lista_tuplas = []
			cpf_nome_colab = cursor.fetchall()

			if len(cpf_nome_colab) == 0:
				inter.pop_up("Atenção", "Não há membros neste laboratório")
				return False

			if tipo=='r':
				for colab in cpf_nome_colab:
					counter = timedelta()
					cursor.execute("SELECT entrada, saida FROM Frequencia WHERE entrada LIKE ? and cpf = ? and saida IS NOT NULL",((date+'%', colab[0])))
					for frequencia in cursor.fetchall():
						entrada = retorna_objeto_date(frequencia[0])
						saida = retorna_objeto_date(frequencia[1])
						counter += (saida-entrada)

					tupla = (colab[1], str(counter))
					lista_tuplas.append(tupla)

			elif tipo=='d':
				dias = ["01","02","03","04","05","06","07","08","09","10","11","12","13","14","15","16","17","18","19","20"
				,"21","22","23","24","25","26","27","28","29","30","31"]
				
				for dia in dias:
					data_com_dia = date+"/"+ dia
					linha = dia+" de "+mes+" de "+ano
					if laboratorio != "Não Ativos":
						cursor.execute('''SELECT Nome, entrada, saida
										  FROM Colaborador as C, Frequencia as F
										  WHERE C.Lab = ? and C.cpf == F.cpf and  C.Status != "Não Ativo" and F.entrada LIKE ? and F.saida IS NOT NULL 
										  ORDER BY Nome''',(laboratorio, data_com_dia+"%"))
					else:
						cursor.execute('''SELECT Nome, entrada, saida
										  FROM Colaborador as C, Frequencia as F
										  WHERE C.cpf == F.cpf and C.Status = "Não Ativo" and F.entrada LIKE ? and F.saida IS NOT NULL 
										  ORDER BY Nome''',(data_com_dia+"%",))
					frequencias = cursor.fetchall()
					
					if frequencias:		
						lista_tuplas.append(("__________________",linha,"__________________"))

						for frequencia in frequencias:
							tupla = (frequencia[0], frequencia[1],frequencia[2])
							lista_tuplas.append(tupla)
				
			return lista_tuplas
		except:
			return False

	else:
		inter.pop_up("ERROR", "Laboratorio inválido!")
		return False

def validar_login(login, senha, event=None):
	
	if (login.get() != "") and (senha.get() != ""):
		try:
			colab = retorna_user(login.get())

			if (colab.senha == senha.get()) and (colab.status == "Ativo") and (colab.funcao != "Pesquisador"):
				return colab

			else:
				inter.pop_up("ERROR", "Login ou Senha Inválida")
				return None

		except Exception as e:
			inter.pop_up("ERROR", "Login ou Senha Inválida")
			return None

	else:
		inter.pop_up("ERROR", "Login ou Senha Inválida")
		return None


def excluir_colaborador(cpf, lab):
	try:
		cursor = con.cursor
		#Seleciona idDigital
		cursor.execute("SELECT idDigital FROM Digital WHERE cpf='%s'"%cpf)

		idDigital = cursor.fetchone()

		if idDigital != None:
			idDigital = idDigital[0]
			#Apaga digital do banco
			cursor.execute("DELETE FROM Digital WHERE cpf= '%s'"%cpf)
			#Desativa o Colaborador
			cursor.execute("UPDATE Colaborador SET Status = (?) WHERE cpf = (?)", ("Não Ativo", cpf))
			#Setta data de Desligamento 
			cursor.execute("UPDATE Colaborador SET DtDesligamento = (?) WHERE cpf = (?)", (retorna_data(), cpf))
			try:
				#Disponibiliza id do colaborador apagado
				cursor.execute('''UPDATE IdDisponivel SET disponivel = (?) WHERE idDigital = (?)''', (0, idDigital))
			except:
				pass

			con.conexao.commit()

		else:
			#Desativa o Colaborador
			cursor.execute("UPDATE Colaborador SET Status = 'Não Ativo' WHERE cpf = (?)", (cpf,))
			con.conexao.commit()
		return True
	except Exception as e:
		print (e)
		inter.pop_up("ERROR", "Não foi possível remover o colaborador.")
		return False


def excluir_lab(nome, sigla):
	lista_colab = retorna_lista_colab(nome)
	if len(lista_colab) == 0:
		try:
			cursor = con.cursor
			cursor.execute("DELETE FROM Laboratorio WHERE Nome = '%s'"%nome)
			con.conexao.commit()
			return True
		except:
			inter.pop_up("ERROR", "Não foi possível remover o laboratório")
	else:
		inter.pop_up("Error", "O laboratório deve estar vazio para ser removido")
		return False
		
def validar_consulta(lab):
	if lab.get() != "" and lab.get() != "*Selecione o laboratório*":
		return True
	else:
		inter.pop_up("ERROR", "Consulta Inválida")
		return False

def validar_consulta_2(nome_colab, lab):
	try:
		colab = retorna_colab(nome_colab, lab)
		return True
	except:
		inter.pop_up("ERROR", "Consulta Inválida")
		return False

def retorna_data_sem_hora(data_str):
	split1 = data_str.split(" ")
	return split1[0]

def retorna_dia(data_str):
	split1 = data_str.split(" ")
	split2 = split1[0]
	data = split2.split("/")
	return data[2]

def retorna_hora(data_str):
	split1 = data_str.split(" ")
	return split1[1]



def gerar_relatorio_resumido(tuplas, lab, mes, ano):
	cursor = con.cursor
	cursor.execute("SELECT Sigla FROM Laboratorio WHERE Nome=(?)",(lab,))
	sigla = cursor.fetchone()[0]

	file_name = "Relatório resumido-"+sigla+"-"+mes+"-"+ano
	file_path = asksaveasfilename(title = "Select file", initialfile=file_name, filetypes = (("Arquivos Excel","*.xlsx"),))
	file_path = file_path+".xlsx"

	alignment_left = Alignment(horizontal='left')
	alignment_right = Alignment(horizontal='right')
	alignment_center = Alignment(horizontal='center')

	font_normal = Font(name='Arial')
	font_bold = Font(name='Arial',bold=True)

	#Criando workbook
	wb = Workbook()

	ws_resumido = wb.active
	ws_resumido.title = "Histórico resumido"

	#inserindo logo 
	img = Image(r"\\LSEHOST\Documents\SCF\imagens\hub.png")
	ws_resumido.add_image(img,'A1')

	#Formatando células para data de emissão do relatório
	ws_resumido.merge_cells('F2:G2')
	ws_resumido.merge_cells('H2:I2')

	cell_emissao = ws_resumido['F2']
	cell_data = ws_resumido['H2']

	cell_emissao.alignment, cell_emissao.font = alignment_right, font_normal
	cell_data.alignment, cell_data.font = alignment_left, font_normal

	now = datetime.now()
	if now.month < 10:
		mes_emissao = "0"+str(now.month)
	else:
		mes_emissao = str(now.month)

	ws_resumido['F2'] = "Data de emissão:"
	ws_resumido['H2'] = str(now.day)+"/"+mes_emissao+"/"+str(now.year)

	#Formatando células para nome do laboratorio
	ws_resumido.merge_cells('A7:I7')
	cell_lab = ws_resumido['A7']
	cell_lab.font, cell_lab.alignment = font_normal, alignment_center
	ws_resumido['A7'] = lab

	#Formatando células mes e ano
	cell_mes_str = ws_resumido['A9']
	cell_mes = ws_resumido['B9']

	cell_ano_str = ws_resumido['H9']
	cell_ano = ws_resumido['I9']

	cell_mes_str.font, cell_mes.font, cell_ano_str.font, cell_ano.font = font_normal, font_normal, font_normal, font_normal
	cell_mes_str.alignment, cell_ano_str.alignment = alignment_right, alignment_right
	cell_mes.alignment, cell_ano.alignment = alignment_left, alignment_left


	ws_resumido['A9'] = "MÊS:"
	ws_resumido['B9'] = mes
	ws_resumido['H9'] = "ANO:"
	ws_resumido['I9'] = ano #Puxar do banco

	# Nome e qtd horas

	ws_resumido.merge_cells('G11:I11')

	cell_nome = ws_resumido['A11']
	cell_qtd = ws_resumido['G11']

	cell_nome.font, cell_qtd.font = font_bold, font_bold
	cell_nome.alignment, cell_qtd.alignment = alignment_left, alignment_center

	ws_resumido['A11'] = "NOME"
	ws_resumido['G11'] = "QTD HORAS POR MÊS"

	row_base = 12

	for colaborador in tuplas:
		row_name = 'A' + str(row_base)
		row_time = 'H' + str(row_base)
		ws_resumido[row_name] = colaborador[0]
		ws_resumido[row_time] = colaborador[1]
		ws_resumido[row_time].alignment = alignment_center
		row_base+=1
	
	wb.template = False
	wb.save(file_path)

def gerar_relatorio_detalhado(lab, mes, ano):
	dict_mes = {'Janeiro': '01','Fevereiro': '02','Março': '03','Abril': '04','Maio': '05','Junho': '06',
						'Julho': '07','Agosto': '08','Setembro': '09','Outubro': '10','Novembro': '11','Dezembro': '12'}
	cursor = con.cursor
	cursor.execute("SELECT Sigla FROM Laboratorio WHERE Nome=(?)",(lab,))
	sigla = cursor.fetchone()[0]

	file_name = "Relatório detalhado-"+sigla+"-"+mes+"-"+ano
	file_path = asksaveasfilename(title = "Select file", initialfile=file_name, filetypes = (("Arquivos Excel","*.xlsx"),))
	file_path = file_path+".xlsx"

	alignment_left = Alignment(horizontal='left')
	alignment_right = Alignment(horizontal='right')
	alignment_center = Alignment(horizontal='center')

	font_normal = Font(name='Arial')
	font_bold = Font(name='Arial',bold=True)

	lista_colab = retorna_lista_colab(lab)

	#Criando workbook
	wb = Workbook()

	ws_array = []
	ws_first = wb.active
	ws_first.title = lista_colab[0]
	ws_dict = {}
	ws_dict[lista_colab[0]] = ws_first
	for i in range(1,len(lista_colab)):
		ws_temp = wb.create_sheet(lista_colab[i])
		ws_dict[lista_colab[i]] = ws_temp

	for colaborador in lista_colab:
		#inserindo logo
		colab = retorna_colab(colaborador, lab)
		current_sheet = ws_dict[colaborador] 
		img = Image('imagens/hub.png')
		current_sheet.add_image(img,'A1')

			#Formatando células para data de emissão do relatório
		current_sheet.merge_cells('F2:G2')
		current_sheet.merge_cells('H2:I2')

		cell_emissao = current_sheet['F2']
		cell_data = current_sheet['H2']

		cell_emissao.alignment, cell_emissao.font = alignment_right, font_normal
		cell_data.alignment, cell_data.font = alignment_left, font_normal

		now = datetime.now()
		if now.month < 10:
			mes_emissao = "0"+str(now.month)
		else:
			mes_emissao = str(now.month)

		current_sheet['F2'] = "Data de emissão:"
		current_sheet['H2'] = str(now.day)+"/"+mes_emissao+"/"+str(now.year)

		name_cell = current_sheet['A8']
		name_cell.font = font_normal
		current_sheet['A8'] = "NOME: "+colaborador 

		lab_cell = current_sheet['A9']
		lab_cell.font = font_normal
		current_sheet['A9'] = "LABORATÓRIO: "+lab

		funcao_cell = current_sheet['A10']
		funcao_cell.font = font_normal
		current_sheet['A10'] = "FUNÇÃO: "+colab.funcao

		cell_mes_str = current_sheet['A12']

		cell_ano_str = current_sheet['H12']
		cell_ano = current_sheet['I12']

		cell_mes_str.font, cell_ano_str.font, cell_ano.font = font_normal, font_normal, font_normal
		cell_mes_str.alignment, cell_ano_str.alignment = alignment_left, alignment_right
		cell_ano.alignment = alignment_left

		current_sheet.merge_cells('A12:B12')

		current_sheet['A12'] = "MÊS: "+mes
		current_sheet['H12'] = "ANO:"
		current_sheet['I12'] = ano

		cell_day = current_sheet['A14']
		cell_day.font, cell_day.alignment = font_bold, alignment_center
		current_sheet['A14'] = "DIA"

		current_sheet.merge_cells('C14:D14')
		current_sheet.merge_cells('H14:I14')


		cell_ingress = current_sheet['C14']
		cell_ingress.font, cell_ingress.alignment = font_bold, alignment_center
		current_sheet['C14'] = "ENTRADA"

		cell_exit = current_sheet['F14']
		cell_exit.font, cell_exit.alignment = font_bold, alignment_center
		current_sheet['F14'] = "SAÍDA"

		cell_cont = current_sheet['H14']
		cell_cont.font, cell_cont.alignment = font_bold, alignment_center
		current_sheet['H14'] = "ACUMULADO"

		data = ano+"/"+dict_mes[mes]
		if lab != "Não Ativos":
			cursor.execute('''SELECT entrada, saida
							  FROM Colaborador as C, Frequencia as F
							  WHERE C.cpf == F.cpf and C.Nome = ? and C.Lab = ? and F.entrada LIKE ? and F.saida IS NOT NULL 
							  ORDER BY Nome''',(colaborador, lab, data+"%"))
		else:
			cursor.execute('''SELECT Nome, entrada, saida
							  FROM Colaborador as C, Frequencia as F
							  WHERE C.Nome = ? and C.cpf == F.cpf and C.Status = "Não Ativo" and F.entrada LIKE ? and F.saida IS NOT NULL 
							  ORDER BY entrada''',(colaborador, data_com_dia+"%"))
		frequencias = cursor.fetchall()
		

		array_frequencias = []
		array_size = len(frequencias)
		if array_size > 0:
			current_frequencia = retorna_data_sem_hora(frequencias[0][0])
			flag = True
			i = 0
			while(flag):
				frequencia_array = []
				while(flag == True and current_frequencia == retorna_data_sem_hora(frequencias[i][0])):
					frequencia_array.append((frequencias[i][0],frequencias[i][1]))
					i+=1
					if i == array_size:
						flag = False
				if flag==True:
					current_frequencia = retorna_data_sem_hora(frequencias[i][0])
				array_frequencias.append(frequencia_array)
		else:
			pass
		

		row_base = 15
		for day in array_frequencias:
			contador = timedelta()
			for ingress in day:
				merge_str_ingress = 'C'+str(row_base)+":"+'D'+str(row_base)
				merge_str_count = 'H'+str(row_base)+":"+'I'+str(row_base)
				current_sheet.merge_cells(merge_str_ingress)
				current_sheet.merge_cells(merge_str_count)

				row_day_str = 'A' + str(row_base)
				row_ingress_str = 'C' + str(row_base)
				row_exit_str = 'F' + str(row_base)
				row_count_str = 'H' + str(row_base)

				row_day = current_sheet[row_day_str]
				row_ingress = current_sheet[row_ingress_str]
				row_exit = current_sheet[row_exit_str]
				row_count = current_sheet[row_count_str]

				row_day.font, row_ingress.font, row_exit.font, row_count.font = font_normal, font_normal, font_normal, font_normal
				row_ingress.alignment, row_exit.alignment, row_count.alignment, row_day.alignment = alignment_center, alignment_center, alignment_center, alignment_center



				current_sheet[row_day_str] = retorna_dia(ingress[0])
				current_sheet[row_ingress_str] = retorna_hora(ingress[0])
				current_sheet[row_exit_str] = retorna_hora(ingress[1])
				delta = timedelta()
				delta = retorna_objeto_date(ingress[1])-retorna_objeto_date(ingress[0])
				contador = contador + delta
				current_sheet[row_count_str] = delta
				row_base+=1

			merge_total = 'H'+str(row_base)+":"+'I'+str(row_base)
			current_sheet.merge_cells(merge_total)

			total_str = 'A'+str(row_base)
			total_num_str = 'H'+str(row_base)
			total_cell = current_sheet[total_str]
			total_num_cell = current_sheet[total_num_str]
			total_cell.font, total_cell.alignment, total_num_cell.font, total_num_cell.alignment = font_bold, alignment_left, font_bold, alignment_center
			current_sheet[total_str]="TOTAL"
			current_sheet[total_num_str] = contador
			row_base +=2

	wb.template = False
	wb.save(file_path)


class McListBox(object):
	def __init__(self, container_o, header, lista):
		self.tree = None
		self._setup_widgets(container_o, header)
		self._build_tree(header, lista)

	def _setup_widgets(self,container_o, header):
		self.container = ttk.Frame(container_o)
		self.container.pack(fill='both', expand=True)
		self.tree = ttk.Treeview(columns=header, show="headings")
		vsb = ttk.Scrollbar(orient="vertical", command=self.tree.yview)
		hsb = ttk.Scrollbar(orient="horizontal", command=self.tree.xview)
		self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
		self.tree.grid(column=0, row=0, sticky='nsew', in_=self.container)
		vsb.grid(column=1, row=0, sticky='ns', in_=self.container)
		hsb.grid(column=0, row=1, sticky='ew', in_=self.container)
		self.container.grid_columnconfigure(0, weight=1)
		self.container.grid_rowconfigure(0, weight=1)
	def _build_tree(self, header, lista):
		for col in header:
			self.tree.heading(col, text=col.title(), command=lambda c=col: sortby(self.tree, c, 0))
			self.tree.column(col, width=tkFont.Font().measure(col.title()))
		for item in lista:
			self.tree.insert('', 'end', values=item)
			for ix, val in enumerate(item):
				col_w = tkFont.Font().measure(val)
				if self.tree.column(header[ix],width=None)<col_w:
					self.tree.column(header[ix], width=col_w)

def sortby(tree, col, descending):
	data = [(tree.set(child, col), child) \
		for child in tree.get_children('')]
	data.sort(reverse=descending)
	for ix, item in enumerate(data):
		tree.move(item[1], '', ix)
	tree.heading(col, command=lambda col=col: sortby(tree, col, \
		int(not descending)))
